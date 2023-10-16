#!/usr/bin/env python
# coding: utf-8

# ## XLS CVS teste
# 
# 
# 

# # Importação de bibliotecas

# # Variáveis parâmetros

# In[1]:


SystemPath = 'TI/XLS_CVS_teste'
FileName = 'XLS_CVS_teste'
DatePath = '2023/10/16/07/00'


# In[2]:


from pyspark.sql.functions import split, col, row_number, expr, row_number, current_timestamp
from pyspark.sql import SparkSession

from pyspark.sql.window import Window

from pyspark import HiveContext
from pyspark.sql import Row, functions as F, types as T
from pyspark.sql.window import Window

from pyspark.sql.functions import col, lit, split, input_file_name, substring, regexp_replace, trim, when, add_months, to_date, concat, coalesce, months_between, max, sum, count, avg

from datetime import datetime, timedelta

from functools import reduce


# # Áreas de Dados

# In[3]:


LZ_path = 'abfss://landing-zone@romagnoledatalake1.dfs.core.windows.net/'+ SystemPath +'/'+ DatePath +'/' + FileName +'.avro'
CZ_path = 'abfss://consume-zone@romagnoledatalake1.dfs.core.windows.net/'+ SystemPath


# # Leitura do arquivo CSV
# # Estrutura original: 
# #PRODUTO;ANO;REGIAO;ESTADO;Jan;Fev;Mar;Abr;Mai;Jun;Jul;Ago;Set;Out;Nov;Dez;TOTAL
# #GASOLINA C (m3);2000;REGIÃO NORTE;RONDÔNIA;136073,253;9563,263;11341,229;9369,746;10719,983;11165,968;12312,451;11220,97;12482,281;13591,122;11940,57;11547,576;10818,094
# #GASOLINA C (m3);2000;REGIÃO NORTE;ACRE;3358,346;40001,853;3065,758;3495,29;2946,93;3023,92;3206,93;3612,58;3264,46;3835,74;3676,571;3225,61;3289,718
# 

# In[4]:


df = spark.read.load('abfss://landing-zone@romagnoledatalake1.dfs.core.windows.net/z_Lab/anp_prod_ok.csv', format='csv', header=True, sep=";")
colunas_a_manter = ['product', 'year', 'unit', 'uf']
#realizando pivot das colunas de Meses para Linhar
df_melted = df.select(colunas_a_manter + [expr("stack(12, 'JAN', JAN, 'FEV', FEV, 'MAR', MAR, 'ABR', ABR, 'MAI', MAI, 'JUN', JUN, 'JUL', JUL, 'AGO', AGO, 'SET', SET, 'OUT', OUT, 'NOV', NOV, 'DEZ', DEZ) as (month, volume)")])
#tipagem de campo e geração de campo Mês numérico (para ordenação)
df = df_melted\
    .withColumn('year', col('year').cast('int'))\
    .withColumn('volume',when(col('volume').isNotNull(), col('volume')).otherwise(0).cast('float'))\
    .withColumn('created_at',current_timestamp())\
    .withColumn('month_no',\
        when(col('month') == 'JAN', 1).\
        when(col('month') == 'FEV', 2).\
        when(col('month') == 'MAR', 3).\
        when(col('month') == 'ABR', 4).\
        when(col('month') == 'MAI', 5).\
        when(col('month') == 'JUN', 6).\
        when(col('month') == 'JUL', 7).\
        when(col('month') == 'AGO', 8).\
        when(col('month') == 'SET', 9).\
        when(col('month') == 'OUT', 10).\
        when(col('month') == 'NOV', 11).\
        when(col('month') == 'DEZ', 12).\
        otherwise(0).cast('int'))     
   
#filtro de registros apenas com valores válidos
df = df.filter(df['volume'].cast('string').rlike('^[0-9.]+$'))
#display(df.limit(12)

#gravar dados na Consume Zone
df.write.mode('overwrite').save(CZ_path, format='parquet')


df_pivot = df.groupBy("month", "month_no").pivot("year").agg(sum("volume")).orderBy("month_no")

agg_cols = df_pivot.columns[1:]
rollup_df = df_pivot.rollup().sum()

renamed_df = reduce(
    lambda rollup_df, idx: rollup_df.withColumnRenamed(rollup_df.columns[idx], agg_cols[idx]), 
    range(len(rollup_df.columns)), rollup_df
)

renamed_df = renamed_df.withColumn('month', lit('Total'))

df_pivot.unionByName(
    renamed_df
)

colunas = df_pivot.columns[2:-1]

# Crie uma janela de especificação para a acumulação
window_spec = Window.orderBy()  # Você pode especificar a ordem conforme necessário

# Calcule a soma acumulativa das duas últimas colunas
for coluna in colunas[-2:]:
    df_pivot = df_pivot.withColumn(coluna + '_cumsum', sum(col(coluna)).over(window_spec))

# Calcule a nova coluna acumulativa com base nas duas últimas colunas
df_pivot = df_pivot.withColumn('variacao acumulado 19-20',
                   when((col(colunas[-2] + '_cumsum') + col(colunas[-1] + '_cumsum') == 0), "n/d")
                   .otherwise((col(colunas[-1] + '_cumsum') / col(colunas[-2] + '_cumsum') - 1) * 100))

# Mostre o DataFrame resultante
df_pivot.show()

#display(final_df.limit(12))


# # Resultados agrupados

# In[ ]:


#agrupamento e sumarização de valor, por ESTADO, PRODUTO, ANO e MES
#df_agregado = df.groupBy("uf", "product", "year", "month_no", "month").agg(sum("volume").alias("TOTAL")).orderBy("uf", "product", "year", "month_no")
#display(df_agregado.limit(12))

#agrupamento e sumarização de valor, por ESTADO, ANO e MES
#df_agregado = df.groupBy("uf", "year", "month_no", "month").agg(sum("volume").alias("volume")).orderBy("uf", "year", "month_no")
#display(df_agregado.limit(12))

#agrupamento e sumarização de valor, por PRODUTO, ANO e MES
#df_agregado = df.groupBy("product", "year", "month_no", "month").agg(sum("volume").alias("volume")).orderBy("year", "month_no", "product")
#display(df_agregado.limit(12))


#filtro por Produto = DIESEL e agrupamento e sumarização de valor, por PRODUTO, ANO e MES
#df_diesel = df.filter(col("product").like("%DIESEL%"))
#df_agregado = df_diesel.groupBy("product", "year", "month_no", "month").agg(sum("volume").alias("accum_total")).orderBy("year", "month_no", "product")
#display(df_agregado.limit(12))


# Use a função sum e over para calcular o total acumulado por mês e ano.
#window_spec = Window.partitionBy("MES").orderBy("ANO").rowsBetween(Window.unboundedPreceding, 0)
#df_totalizado = df.withColumn("accum_total", sum("volume").over(window_spec))
#display(df_totalizado.limit(12))


# In[ ]:


get_ipython().run_cell_magic('pyspark', '', "df = spark.read.load('abfss://landing-zone@romagnoledatalake1.dfs.core.windows.net/z_Lab/Relatório CPR091.csv', format='csv'\r\n## If\u202fheader\u202fexists\u202funcomment\u202fline\u202fbelow\r\n, header=True\r\n)\r\ndisplay(df.limit(10))\n")


# In[ ]:


import pandas as pd

df = pd.read_excel('abfss://landing-zone@romagnoledatalake1.dfs.core.windows.net/z_Lab/Relatório CPR091.xls', engine='openpyxl')
display(df.limit(10))


# In[ ]:


import pandas as pd

df = pd.read_excel('abfss://landing-zone@romagnoledatalake1.dfs.core.windows.net/z_Lab/anp_LibreOffice.xlsx', engine='openpyxl')
display(df.limit(10))


# In[ ]:


from colorama import Back, Fore, init
from openpyxl import load_workbook
 
ROWS, COLS = 10, 10
WORKBOOK, WORKSHEET = "abfss://landing-zone@romagnoledatalake1.dfs.core.windows.net/z_Lab/anp.xlsx", "Plan1"
 
wb = load_workbook(WORKBOOK)
ws = wb[WORKSHEET]
 
init()
 
for r in range(1, ROWS + 1):
    print(
        Fore.CYAN if r % 2 else Fore.MAGENTA,
        " | ".join(
            [
                "{:5d}".format(
                    ws.cell(row=r, column=c).value
                )
                for c in range(1, COLS + 1)
            ]
        ),
        Fore.RESET, sep=""
    )

